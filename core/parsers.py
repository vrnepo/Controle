"""
Leitura dos arquivos do banco.

Este módulo é a razão de o sistema ser um backend Python e não um script na
planilha: a fatura do cartão Santander vem em PDF **protegido por senha** e com
layout de duas colunas. Apps Script não tem biblioteca de PDF e a conversão do
Drive falha em arquivo criptografado; aqui é só passar a senha ao pdfplumber.

Contrato de saída — sempre estas duas listas:

    Leitura.linhas  = [{data, descricao, valor, conta, competencia, arquivo}]
    Leitura.resumos = [{conta, competencia, saldo_anterior, despesas, encargos,
                        creditos, pagamentos, total_informado, arquivo}]

Sinal: saída de dinheiro é NEGATIVA. Estorno e crédito, positivos.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import re
from typing import Any, Callable, Dict, List, Optional, Tuple

from core import config
from core.categorizacao import normalizar


class ErroDeLeitura(Exception):
    """Falha que o usuário precisa ler na tela, com o motivo em português."""


class Leitura:
    def __init__(self, formato: str = "", conta: str = ""):
        self.formato = formato
        self.conta = conta
        self.linhas: List[Dict[str, Any]] = []
        self.resumos: List[Dict[str, Any]] = []
        self.avisos: List[str] = []


# --------------------------------------------------------------- utilidades

def num_br(valor: object) -> Optional[float]:
    """'1.234,56' · '-R$ 89,90' · '− 12,00' · '(45,00)' → float."""
    if isinstance(valor, (int, float)):
        return float(valor)
    t = str(valor or "").strip()
    if not t:
        return None
    t = t.replace("−", "-").replace("–", "-").replace("—", "-")
    t = re.sub(r"R\$", "", t, flags=re.I).replace(" ", "").replace("\xa0", "")
    negativo_parenteses = bool(re.match(r"^\(.*\)$", t))
    if negativo_parenteses:
        t = t[1:-1]
    # vírgula é decimal quando vem depois do último ponto
    if "," in t and t.rfind(",") > t.rfind("."):
        t = t.replace(".", "").replace(",", ".")
    else:
        t = t.replace(",", "")
    try:
        n = float(t)
    except ValueError:
        return None
    return -abs(n) if negativo_parenteses else n


def data_br(valor: object, ano_dica: Optional[int] = None) -> Optional[dt.date]:
    if isinstance(valor, dt.datetime):
        return valor.date()
    if isinstance(valor, dt.date):
        return valor
    t = str(valor or "").strip()
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", t)
    if m:
        return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", t)
    if m:
        ano = int(m.group(3))
        if ano < 100:
            ano += 2000
        return dt.date(ano, int(m.group(2)), int(m.group(1)))
    if ano_dica:
        m = re.match(r"^(\d{1,2})/(\d{1,2})$", t)
        if m:
            return dt.date(ano_dica, int(m.group(2)), int(m.group(1)))
        m = re.match(r"^(\d{1,2})\s+([A-Za-z]{3})$", t)
        if m:
            mes = config.MES_ABREV.get(normalizar(m.group(2)))
            if mes:
                return dt.date(ano_dica, mes, int(m.group(1)))
    return None


def primeiro_do_mes(d: dt.date) -> dt.date:
    return dt.date(d.year, d.month, 1)


def competencia_do_nome(nome: str) -> Optional[dt.date]:
    """
    Competência a partir do nome do arquivo. Para cartão ela NÃO pode ser
    deduzida da data da compra: uma compra de 31/07 entra na fatura de agosto.
    """
    m = re.search(r"(\d{4})-(\d{2})-\d{2}", nome)            # Nubank_2026-08-17
    if m:
        return dt.date(int(m.group(1)), int(m.group(2)), 1)
    m = re.search(r"Fatura_?(\d{2})(\d{4})", nome, re.I)     # Fatura_082026
    if m:
        return dt.date(int(m.group(2)), int(m.group(1)), 1)
    m = re.search(r"(\d{2})([A-Za-z]{3})(\d{4})", nome)      # 01AGO2026
    if m:
        mes = config.MES_ABREV.get(normalizar(m.group(2)))
        if mes:
            return dt.date(int(m.group(3)), mes, 1)
    m = re.search(r"(\d{4})[-_](\d{2})", nome)
    if m:
        return dt.date(int(m.group(1)), int(m.group(2)), 1)
    return None


def conta_do_nome(nome: str) -> Optional[str]:
    n = normalizar(nome)
    if re.search(r"NUBANK_\d{4}-\d{2}-\d{2}", n):
        return "Cartão Nubank"
    if re.match(r"^NU_\d+", n) or "NUCONTA" in n:
        return "Conta Nubank"
    if "SANTANDER" in n and "FATURA" in n:
        return "Cartão Santander"
    if re.search(r"FATURA_\d{6}", n):
        return "Cartão Santander"
    if "SANTANDER" in n:
        return "Conta Santander"
    if "NUBANK" in n or n.startswith("NU-") or n.startswith("NU_"):
        return "Cartão Nubank"
    return None


# Linhas que a fatura usa só para fechar saldo e que se anulam entre si.
# Entrar com elas dobraria o gasto do mês. Encargo real (juros, IOF, multa,
# estorno) NÃO está aqui — aquilo é gasto de verdade.
_CONTABEIS = [
    # "PAGAMENTO EM 19 JAN" e "SALDO EM ROTATIVO" são as variantes que o PDF
    # usa e que escaparam do filtro na reimportação de 08/08/2026 — cinco
    # pagamentos e três saldos entraram como lançamento até a auditoria pegar.
    r"^PAGAMENTO RECEBIDO", r"^PAGAMENTO DE FATURA", r"^PAGAMENTO EFETUADO",
    r"^PAGAMENTO EM \d",
    r"^SALDO (EM ATRASO|EM ROTATIVO|RESTANTE|DO ROTATIVO|ANTERIOR|FINANCIADO)",
    r"^CREDITO DE (ATRASO|ROTATIVO)", r"^ENCERRAMENTO DE DIVIDA",
    r"^JUROS DE DIVIDA ENCERRADA", r"^TOTAL ",
]
_CONTABEIS_RX = [re.compile(p) for p in _CONTABEIS]


def linha_contabil(descricao: str) -> bool:
    d = normalizar(descricao)
    return any(rx.search(d) for rx in _CONTABEIS_RX)


# ----------------------------------------------------------- roteamento

def ler(nome: str, dados: bytes, senha_pdf: str = "") -> Leitura:
    """Descobre o formato pelo conteúdo e delega. `dados` é o arquivo cru."""
    ext = (re.search(r"\.([A-Za-z0-9]+)$", nome) or [None, ""])[1].lower() \
        if re.search(r"\.([A-Za-z0-9]+)$", nome) else ""

    if ext in ("csv", "txt"):
        return _ler_csv(nome, dados)
    if ext in ("ofx", "qfx"):
        return _ler_ofx(nome, dados)
    if ext in ("xlsx", "xlsm"):
        return _ler_xlsx(nome, dados)
    if ext == "pdf":
        return _ler_pdf(nome, dados, senha_pdf)
    raise ErroDeLeitura("Extensão não suportada: .%s (use csv, xlsx, ofx ou pdf)" % ext)


def _matriz_csv(dados: bytes) -> List[List[str]]:
    texto = dados.decode("utf-8-sig", errors="replace")
    primeira = texto.split("\n", 1)[0]
    sep = ","
    if primeira.count(";") > primeira.count(","):
        sep = ";"
    elif primeira.count("\t") > primeira.count(","):
        sep = "\t"
    return [linha for linha in csv.reader(io.StringIO(texto), delimiter=sep)]


def _cabecalho(matriz: List[List[str]]) -> List[str]:
    return [normalizar(x) for x in (matriz[0] if matriz else [])]


def _indice(cab: List[str], nomes: List[str]) -> int:
    for n in nomes:
        alvo = normalizar(n)
        if alvo in cab:
            return cab.index(alvo)
    return -1


def _ler_csv(nome: str, dados: bytes) -> Leitura:
    matriz = _matriz_csv(dados)
    if len(matriz) < 2:
        raise ErroDeLeitura("Arquivo vazio ou só com cabeçalho.")
    cab = _cabecalho(matriz)

    if "SALDO_ANTERIOR" in cab or "TOTAL_INFORMADO" in cab:
        return _csv_resumo(nome, matriz, cab)
    if "CONTA" in cab and "COMPETENCIA" in cab and "VALOR" in cab:
        return _csv_normalizado(nome, matriz, cab)
    if "DATE" in cab and "TITLE" in cab and "AMOUNT" in cab:
        return _csv_nubank_fatura(nome, matriz)
    if "IDENTIFICADOR" in cab or ("DATA" in cab and "VALOR" in cab):
        return _csv_extrato(nome, matriz, cab)
    raise ErroDeLeitura("Não reconheci o formato. Cabeçalho lido: " + " | ".join(cab))


def _csv_nubank_fatura(nome: str, matriz: List[List[str]]) -> Leitura:
    comp = competencia_do_nome(nome)
    if not comp:
        raise ErroDeLeitura(
            "Não consegui deduzir a competência de '%s'. Renomeie no padrão "
            "Nubank_AAAA-MM-DD.csv (a data de vencimento da fatura)." % nome)
    r = Leitura("CSV fatura Nubank", "Cartão Nubank")
    for linha in matriz[1:]:
        if len(linha) < 3 or not (linha[0] or "").strip():
            continue
        d = data_br(linha[0], comp.year)
        v = num_br(linha[2])
        if d is None or v is None:
            continue
        desc = (linha[1] or "").strip()
        if linha_contabil(desc):
            continue
        r.linhas.append({"data": d, "descricao": desc, "valor": -v,
                         "conta": "Cartão Nubank", "competencia": comp, "arquivo": nome})
    return r


def _csv_extrato(nome: str, matriz: List[List[str]], cab: List[str]) -> Leitura:
    conta = conta_do_nome(nome)
    if not conta:
        raise ErroDeLeitura(
            "Não sei a que conta '%s' pertence. Inclua 'Nubank' ou 'Santander' "
            "no nome do arquivo." % nome)
    iD = _indice(cab, ["DATA", "DATE"])
    iV = _indice(cab, ["VALOR", "AMOUNT"])
    iT = _indice(cab, ["DESCRICAO", "TITLE", "HISTORICO", "LANCAMENTO"])
    r = Leitura("CSV extrato", conta)
    cartao = conta.startswith("Cartão")
    comp_fixa = competencia_do_nome(nome) if cartao else None
    for linha in matriz[1:]:
        if iD >= len(linha) or iV >= len(linha):
            continue
        d = data_br(linha[iD])
        v = num_br(linha[iV])
        if d is None or v is None:
            continue
        desc = (linha[iT] or "").strip() if 0 <= iT < len(linha) else "(sem descrição)"
        if cartao and linha_contabil(desc):
            continue
        r.linhas.append({"data": d, "descricao": desc,
                         "valor": (-v if cartao else v), "conta": conta,
                         "competencia": comp_fixa or primeiro_do_mes(d), "arquivo": nome})
    return r


def _csv_normalizado(nome: str, matriz: List[List[str]], cab: List[str]) -> Leitura:
    iD, iT = _indice(cab, ["DATA"]), _indice(cab, ["DESCRICAO"])
    iV, iC = _indice(cab, ["VALOR"]), _indice(cab, ["CONTA"])
    iK, iA = _indice(cab, ["COMPETENCIA"]), _indice(cab, ["ARQUIVO"])
    r = Leitura("CSV normalizado", "")
    for linha in matriz[1:]:
        if max(iD, iT, iV, iC) >= len(linha):
            continue
        d, v = data_br(linha[iD]), num_br(linha[iV])
        conta = (linha[iC] or "").strip()
        if d is None or v is None or not conta:
            continue
        comp = None
        if 0 <= iK < len(linha):
            bruto = (linha[iK] or "").strip()
            comp = data_br(bruto + "-01" if len(bruto) == 7 else bruto)
        r.linhas.append({
            "data": d, "descricao": (linha[iT] or "").strip(), "valor": v, "conta": conta,
            "competencia": primeiro_do_mes(comp or d),
            "arquivo": (linha[iA] if 0 <= iA < len(linha) and linha[iA] else nome)})
    return r


def _csv_resumo(nome: str, matriz: List[List[str]], cab: List[str]) -> Leitura:
    r = Leitura("CSV resumo de fatura", "")
    iC, iK = _indice(cab, ["CONTA"]), _indice(cab, ["COMPETENCIA"])

    def campo(linha: List[str], nomes: List[str]) -> float:
        i = _indice(cab, nomes)
        return (num_br(linha[i]) or 0.0) if 0 <= i < len(linha) else 0.0

    for linha in matriz[1:]:
        if max(iC, iK) >= len(linha):
            continue
        conta = (linha[iC] or "").strip()
        bruto = (linha[iK] or "").strip()
        comp = data_br(bruto + "-01" if len(bruto) == 7 else bruto)
        if not conta or comp is None:
            continue
        r.resumos.append({
            "conta": conta, "competencia": primeiro_do_mes(comp),
            "saldo_anterior": campo(linha, ["SALDO_ANTERIOR"]),
            "despesas": campo(linha, ["DESPESAS", "DESPESAS_BR"]) + campo(linha, ["DESPESAS_EXT"]),
            "encargos": campo(linha, ["ENCARGOS"]),
            "creditos": campo(linha, ["CREDITOS"]),
            "pagamentos": campo(linha, ["PAGAMENTOS"]),
            "total_informado": campo(linha, ["TOTAL_INFORMADO", "TOTAL"]),
            "arquivo": nome})
    return r


def _ler_xlsx(nome: str, dados: bytes) -> Leitura:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(dados), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    matriz: List[List[str]] = []
    for linha in ws.iter_rows(values_only=True):
        if linha is None:
            continue
        matriz.append(["" if c is None else c for c in linha])
    wb.close()
    if len(matriz) < 2:
        raise ErroDeLeitura("Planilha vazia.")
    # reaproveita os leitores de CSV: eles só olham a matriz
    cab = _cabecalho(matriz)
    if "SALDO_ANTERIOR" in cab or "TOTAL_INFORMADO" in cab:
        return _csv_resumo(nome, matriz, cab)
    if "CONTA" in cab and "COMPETENCIA" in cab and "VALOR" in cab:
        return _csv_normalizado(nome, matriz, cab)
    if "DATE" in cab and "TITLE" in cab and "AMOUNT" in cab:
        return _csv_nubank_fatura(nome, matriz)
    return _csv_extrato(nome, matriz, cab)


def _ler_ofx(nome: str, dados: bytes) -> Leitura:
    texto = dados.decode("latin-1", errors="replace")
    conta = conta_do_nome(nome) or "Conta Santander"
    r = Leitura("OFX", conta)
    for bloco in re.split(r"<STMTTRN>", texto, flags=re.I)[1:]:
        mdt = re.search(r"<DTPOSTED>\s*(\d{8})", bloco, re.I)
        mvl = re.search(r"<TRNAMT>\s*(-?[\d.,]+)", bloco, re.I)
        if not mdt or not mvl:
            continue
        v = num_br(mvl.group(1))
        if v is None:
            continue
        bruto = mdt.group(1)
        d = dt.date(int(bruto[:4]), int(bruto[4:6]), int(bruto[6:8]))
        nm = (re.search(r"<NAME>\s*([^<\r\n]*)", bloco, re.I) or [None, ""])[1]
        me = (re.search(r"<MEMO>\s*([^<\r\n]*)", bloco, re.I) or [None, ""])[1]
        r.linhas.append({"data": d, "descricao": (nm + " " + me).strip(), "valor": v,
                         "conta": conta, "competencia": primeiro_do_mes(d), "arquivo": nome})
    if not r.linhas:
        raise ErroDeLeitura("Nenhuma transação encontrada no OFX.")
    return r


# ------------------------------------------------------------------- PDF

def _paginas_pdf(dados: bytes, senha: str) -> Tuple[List[str], List[str]]:
    """
    Devolve (textos_inteiros, metades).

    As metades existem por causa do Santander: o extrato da fatura vem em DUAS
    COLUNAS na mesma página, e extrair a página inteira mistura as linhas das
    duas — datas de uma coluna colam em valores da outra. Recortar cada metade
    antes de extrair resolve.
    """
    import pdfplumber

    try:
        pdf = pdfplumber.open(io.BytesIO(dados), password=senha or "")
    except Exception as erro:
        if re.search(r"password|encrypt|decrypt", str(erro), re.I):
            raise ErroDeLeitura(
                "PDF protegido por senha e a senha configurada não abriu. "
                "Confira SENHA_PDF_SANTANDER nas variáveis de ambiente.")
        raise ErroDeLeitura("Não consegui abrir o PDF: %s" % erro)

    inteiros, metades = [], []
    with pdf:
        for p in pdf.pages:
            inteiros.append(p.extract_text() or "")
            largura, altura = p.width, p.height
            metades.append(p.crop((0, 0, largura / 2, altura)).extract_text() or "")
            metades.append(p.crop((largura / 2, 0, largura, altura)).extract_text() or "")
    return inteiros, metades


def _ler_pdf(nome: str, dados: bytes, senha: str) -> Leitura:
    conta = conta_do_nome(nome)
    inteiros, metades = _paginas_pdf(dados, senha)
    texto = "\n".join(inteiros)

    if "Saldo Anterior" in texto or re.search(r"Total Despesas/D[ée]bitos", texto):
        return _pdf_santander_fatura(nome, texto, metades)
    if conta == "Conta Nubank" or re.search(r"NuConta|Nu Financeira|Extrato da conta", texto, re.I):
        return _pdf_nubank_extrato(nome, texto)
    return _pdf_nubank_fatura(nome, texto)


# Rótulos de encargo do resumo do Santander.
#
# A lista antiga tinha só "Juros Remuneratórios" — e na fatura de jul/2026 o
# banco escreveu "Juros de Crédito Rotativo". R$ 726,52 ficaram fora da planilha
# sem nenhum aviso, porque o parser não achou o rótulo que esperava e seguiu
# adiante. Toda variante que o Santander usa entra aqui.
ENCARGOS_SANTANDER = [
    "Juros Remuneratórios",
    "Juros de Crédito Rotativo",
    "Juros de Parcelamento",
    "Juros de Mora",
    "IOF",
    "Multa por Atraso",
    "Encargos por Atraso",
]

_LINHA_SANTANDER = re.compile(
    r"^(?:[23]\s+)?(\d{2})/(\d{2})\s+(.+?)\s+(-?[\d.]+,\d{2})(?:\s+(-?[\d.]+,\d{2}))?$")
_IOF_EXTERIOR = re.compile(r"^(IOF DESPESA NO EXTERIOR)\s+(-?[\d.]+,\d{2})$")
_SECOES_SANTANDER = [("Pagamento e Demais Créditos", "creditos"),
                     ("Parcelamentos", "parcelamentos"),
                     ("Despesas", "despesas")]


def _pdf_santander_fatura(nome: str, texto: str, metades: List[str]) -> Leitura:
    comp = competencia_do_nome(nome)
    if not comp:
        raise ErroDeLeitura(
            "Não consegui deduzir a competência de '%s' (esperava algo como "
            "Fatura_082026 no nome)." % nome)
    ano, mes = comp.year, comp.month
    r = Leitura("PDF fatura Santander", "Cartão Santander")

    def campo(rotulo: str, sinal: str = r"\+") -> float:
        m = re.search(r"\(" + sinal + r"\)\s*" + re.escape(rotulo) + r"\s+(-?[\d.]+,\d{2})",
                      texto)
        return num_br(m.group(1)) or 0.0 if m else 0.0

    m = re.search(r"Saldo Anterior\s+(-?[\d.]+,\d{2})", texto)
    saldo = num_br(m.group(1)) or 0.0 if m else 0.0

    encargos_detalhe: List[Tuple[str, float]] = []
    for rotulo in ENCARGOS_SANTANDER:
        v = campo(rotulo)
        if abs(v) > 0.005:
            encargos_detalhe.append((rotulo, v))
    encargos = sum(v for _, v in encargos_detalhe)

    despesas = campo("Total Despesas/Débitos no Brasil")
    m = re.search(r"\(\+\)\s*Total Despesas/Débitos no Exterior\s+(-?[\d.]+,\d{2})", texto)
    if m:
        despesas += num_br(m.group(1)) or 0.0
    creditos = campo("Total de créditos", sinal="-")
    pagamentos = campo("Total de pagamentos", sinal="-")
    m = re.search(r"Pagamento Total\s+R\$\s*([\d.]+,\d{2})", texto)
    total_informado = num_br(m.group(1)) or 0.0 if m else 0.0

    r.resumos.append({
        "conta": "Cartão Santander", "competencia": comp, "saldo_anterior": saldo,
        "despesas": despesas, "encargos": encargos, "creditos": creditos,
        "pagamentos": pagamentos, "total_informado": total_informado, "arquivo": nome})

    # --- itens, seção por seção.
    #
    # `secao` NÃO se reinicia a cada metade: a tabela começa numa coluna e
    # continua na seguinte sem repetir o título. Reiniciar aqui descartava tudo
    # o que vinha depois da primeira coluna (em jul/2026 sobravam R$ 4.254 de
    # R$ 8.477).
    secao = None
    soma_itens = 0.0
    for metade in metades:
        for bruta in metade.split("\n"):
            s = bruta.strip()
            for marcador, nome_secao in _SECOES_SANTANDER:
                if s.startswith(marcador):
                    secao = nome_secao
                    break
            if secao is None:
                continue
            mi = _IOF_EXTERIOR.match(s)
            if mi:
                v = num_br(mi.group(2)) or 0.0
                r.linhas.append({"data": dt.date(ano, mes, 1),
                                 "descricao": "IOF despesa no exterior", "valor": -v,
                                 "conta": "Cartão Santander", "competencia": comp,
                                 "arquivo": nome})
                soma_itens += v
                continue
            mm = _LINHA_SANTANDER.match(s)
            if not mm:
                continue
            dia, mon = int(mm.group(1)), int(mm.group(2))
            desc, valor = mm.group(3).strip(), num_br(mm.group(4)) or 0.0
            if secao == "parcelamentos":
                mp = re.match(r"^(.*?)\s+(\d{2}/\d{2})$", desc)
                if mp:
                    desc = "%s - Parcela %s" % (mp.group(1), mp.group(2))
            if "PAGAMENTO" in desc.upper() or "PGTO" in desc.upper():
                continue
            if abs(valor) < 0.005 or not (1 <= mon <= 12) or not (1 <= dia <= 31):
                continue
            # compra de dezembro numa fatura de janeiro é do ano anterior
            a = ano - 1 if mon > mes else ano
            try:
                data = dt.date(a, mon, dia)
            except ValueError:
                continue
            r.linhas.append({"data": data, "descricao": desc.replace(",", " "),
                             "valor": -valor, "conta": "Cartão Santander",
                             "competencia": comp, "arquivo": nome})
            soma_itens += valor

    for rotulo, v in encargos_detalhe:
        r.linhas.append({"data": dt.date(ano, mes, 1), "descricao": "%s (fatura)" % rotulo,
                         "valor": -v, "conta": "Cartão Santander", "competencia": comp,
                         "arquivo": nome})

    # Diferença residual fica EXPLÍCITA como lançamento, em vez de virar um
    # buraco silencioso: se o PDF tinha linha ilegível, o mês continua fechando
    # com o resumo oficial e a linha diz de onde veio.
    declarado = despesas - creditos
    diferenca = round(declarado - soma_itens, 2)
    if abs(diferenca) > 0.05:
        r.linhas.append({"data": dt.date(ano, mes, 1),
                         "descricao": "Ajuste de importação (linha ilegível no PDF)",
                         "valor": -diferenca, "conta": "Cartão Santander",
                         "competencia": comp, "arquivo": nome})
        r.avisos.append("Ajuste de %s: a soma dos itens ficou %s e o resumo declara %s."
                        % (diferenca, round(soma_itens, 2), round(declarado, 2)))

    # Decisão do usuário (08/08/2026): na fatura do SANTANDER, os lançamentos
    # do mês devem somar exatamente o "(=) Saldo Desta Fatura" do resumo. Por
    # isso, quando há saldo anterior aumentando a fatura, ele entra como
    # lançamento — e os pagamentos recebidos nela também (só nesse caso).
    # As duas linhas são categorizadas como TRANSFERÊNCIA pelas Regras: somam
    # no total da fatura, mas não contam como despesa do mês — o saldo
    # anterior é gasto do mês passado, que já foi contado lá.
    if saldo > 0.005:
        r.linhas.append({"data": dt.date(ano, mes, 1),
                         "descricao": "Saldo anterior da fatura", "valor": -saldo,
                         "conta": "Cartão Santander", "competencia": comp,
                         "arquivo": nome})
        if pagamentos > 0.005:
            r.linhas.append({"data": dt.date(ano, mes, 1),
                             "descricao": "Pagamentos recebidos na fatura",
                             "valor": pagamentos, "conta": "Cartão Santander",
                             "competencia": comp, "arquivo": nome})
    return r


_LINHA_NUBANK = re.compile(
    r"^(\d{1,2})\s+([A-Z]{3})\s+(.+?)\s+(-?[−–]?\s?R\$\s?[\d.]+,\d{2})$")


def _pdf_nubank_fatura(nome: str, texto: str) -> Leitura:
    comp = competencia_do_nome(nome)
    if not comp:
        raise ErroDeLeitura("Não consegui deduzir a competência de '%s'." % nome)
    r = Leitura("PDF fatura Nubank", "Cartão Nubank")
    for bruta in texto.split("\n"):
        s = re.sub(r"^[••]+\s*", "", bruta).strip()
        m = _LINHA_NUBANK.match(s)
        if not m:
            continue
        mes = config.MES_ABREV.get(normalizar(m.group(2)))
        if not mes:
            continue
        v = num_br(m.group(4))
        if v is None:
            continue
        desc = m.group(3).strip()
        # O PDF prefixa cada compra com o final do cartão ("•••• 8751 Loja X");
        # o CSV exportado pelo app e todo o histórico gravam só "Loja X".
        # Manter o prefixo mudaria a chave de deduplicação e a MESMA fatura,
        # importada uma vez por CSV e outra por PDF, entraria duas vezes —
        # visto na prática em 08/08/2026, no diff de jul/2026 (53 linhas iguais
        # que só diferiam pelo prefixo).
        desc = re.sub(r"^[•·*]+\s*\d{4}\s+", "", desc)
        # O PDF escreve 'IOF de "Openai"' com aspas; o CSV e o histórico, sem.
        # Com aspas a chave de dedup muda e a mesma linha entra duas vezes —
        # foram 8 duplicatas na reimportação de 08/08/2026 até tirar isto.
        desc = desc.replace('"', "").replace("“", "").replace("”", "").strip()
        if linha_contabil(desc):
            continue
        ano = comp.year - 1 if mes > comp.month else comp.year
        try:
            data = dt.date(ano, mes, int(m.group(1)))
        except ValueError:
            continue
        r.linhas.append({"data": data, "descricao": desc, "valor": -v,
                         "conta": "Cartão Nubank", "competencia": comp, "arquivo": nome})
    if not r.linhas:
        raise ErroDeLeitura(
            "Não achei lançamentos no PDF da fatura. O app do Nubank exporta a "
            "fatura em CSV — prefira o CSV, que é exato.")

    # RESUMO DA FATURA ATUAL (decisão do usuário, 08/08/2026): a soma dos
    # lançamentos da fatura Nubank deve ser o "Total a pagar". A "Fatura
    # anterior" pode ser NEGATIVA (crédito — jul/26 veio com −258,01), e o
    # próprio Nubank arredonda o total (jul/26: componentes somam 3.260,97 e
    # o app mostra 3.260,98) — por isso a linha de ajuste ao final.
    #
    # A busca é RECORTADA à seção do quadro e ancorada no início da linha: a
    # 1ª versão varria o documento inteiro, casava "Total a pagar" de outras
    # partes do PDF e fabricava ajustes de centenas de reais (dry-run de
    # 08/08/2026). O fim da seção varia por mês: nos meses de rotativo não
    # existe a linha "Pagamento mínimo" — por isso a lista de terminadores.
    inicio = texto.upper().find("RESUMO DA FATURA")
    secao = ""
    if inicio >= 0:
        fim = len(texto)
        for terminador in ("Pagamento mínimo", "O Nubank declara", "PRÓXIMAS FATURAS"):
            achado = texto.find(terminador, inicio)
            if achado > 0:
                fim = min(fim, achado)
        secao = texto[inicio:fim]

    def campo(rotulo: str) -> Optional[float]:
        m = re.search(r"^" + rotulo + r"[^\n]*?(−?-?\s?R\$\s?[\d.]+,\d{2})\s*$",
                      secao, re.M)
        return num_br(m.group(1)) if m else None

    saldo = campo(r"Fatura anterior")
    total_a_pagar = campo(r"Total a pagar")
    if saldo is not None and total_a_pagar is not None:
        compras = campo(r"Total de compras de todos os cart.es") or 0.0
        iof = campo(r"IOF de compras internacionais") or 0.0
        outros = campo(r"Outros lan.amentos") or 0.0
        pagamentos = abs(campo(r"Pagamento recebido") or 0.0)   # o PDF o mostra negativo
        # Meses de rotativo trazem linhas extras no quadro; elas correspondem
        # a lançamentos reais da fatura (Juros/IOF de rotativo, Estorno de
        # juros) e entram nos ENCARGOS do resumo, senão a conciliação veria
        # o "novo" menor do que os itens somam. "Saldo financiado" é
        # intermediário (saldo − pagamento) e fica de fora de propósito.
        juros_fin = campo(r"Juros de financiamento") or 0.0
        iof_fin = campo(r"IOF de financiamento") or 0.0
        estorno_juros = campo(r"Estorno de juros") or 0.0      # vem negativo
        r.resumos.append({
            "conta": "Cartão Nubank", "competencia": comp,
            "saldo_anterior": saldo,
            "despesas": compras + outros,    # "Outros" (estornos etc.) abate aqui
            "encargos": iof + juros_fin + iof_fin + estorno_juros,
            "creditos": 0.0,
            "pagamentos": pagamentos, "total_informado": total_a_pagar,
            "arquivo": nome})
        if abs(saldo) > 0.005:
            r.linhas.append({"data": comp, "descricao": "Saldo anterior da fatura",
                             "valor": -saldo, "conta": "Cartão Nubank",
                             "competencia": comp, "arquivo": nome})
        if pagamentos > 0.005:
            r.linhas.append({"data": comp,
                             "descricao": "Pagamentos recebidos na fatura",
                             "valor": pagamentos, "conta": "Cartão Nubank",
                             "competencia": comp, "arquivo": nome})
        # Trava final: a soma TEM de ser o Total a pagar. Diferença de até 5
        # centavos é arredondamento do próprio Nubank e vira a linha de
        # ajuste; acima disso é item faltando ou leitura errada — aí NÃO se
        # fabrica ajuste: fica o aviso, e a Conciliação aponta o buraco.
        diferenca = round(-total_a_pagar - sum(l["valor"] for l in r.linhas), 2)
        if 0.005 < abs(diferenca) <= 0.05:
            r.linhas.append({"data": comp,
                             "descricao": "Ajuste de arredondamento da fatura",
                             "valor": diferenca, "conta": "Cartão Nubank",
                             "competencia": comp, "arquivo": nome})
        elif abs(diferenca) > 0.05:
            r.avisos.append(
                "A soma dos lançamentos difere do Total a pagar em %s — item "
                "faltando ou leitura errada; nenhum ajuste foi inventado."
                % diferenca)
    return r


_LINHA_NU_EXTRATO = re.compile(
    r"^(\d{1,2})\s+([A-Z]{3})\s+(\d{4})\s+(.+?)\s+(-?[−–]?\s?R\$\s?[\d.]+,\d{2})$")


def _pdf_nubank_extrato(nome: str, texto: str) -> Leitura:
    r = Leitura("PDF extrato Nubank", "Conta Nubank")
    ano_dica = None
    m = re.search(r"(\d{4})", nome)
    if m:
        ano_dica = int(m.group(1))
    for bruta in texto.split("\n"):
        s = bruta.strip()
        m = _LINHA_NU_EXTRATO.match(s)
        if m:
            mes = config.MES_ABREV.get(normalizar(m.group(2)))
            v = num_br(m.group(5))
            if not mes or v is None:
                continue
            try:
                data = dt.date(int(m.group(3)), mes, int(m.group(1)))
            except ValueError:
                continue
            desc = m.group(4).strip()
            # no extrato, saída já vem com sinal negativo no texto
            r.linhas.append({"data": data, "descricao": desc, "valor": v,
                             "conta": "Conta Nubank",
                             "competencia": primeiro_do_mes(data), "arquivo": nome})
    if not r.linhas:
        raise ErroDeLeitura(
            "Não achei lançamentos no extrato em PDF. O Nubank também exporta o "
            "extrato em CSV e em OFX — qualquer um dos dois é mais confiável.")
    return r
