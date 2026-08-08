#!/usr/bin/env python3
"""
Carga inicial: leva o histórico de 2026 que já existe no
"Controle Financeiro 2.0.xlsx" para o banco.

    DATABASE_URL='postgresql://...' python3 scripts/migrar_da_planilha.py \
        "/caminho/Controle Financeiro 2.0.xlsx"

São 1.874 lançamentos de jan a ago/2026, já categorizados e conferidos contra os
PDFs. Reimportar tudo pelos arquivos originais daria quase o mesmo resultado,
mas perderia as correções manuais de categoria feitas ao longo do caminho — e
duas faturas (Santander de mar/26 e Nubank de abr/26) não estão na pasta, então
esses meses só existem aqui.

A carga é idempotente: usa a MESMA chave de deduplicação da importação normal,
então rodar duas vezes não duplica nada.

Junto vão os 7 resumos oficiais das faturas Santander de 2026, extraídos dos
PDFs e conferidos ao centavo contra o total que o app do banco mostra.
"""

from __future__ import annotations

import datetime as dt
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Carrega o .env da raiz do projeto ANTES de importar core.* — sem isto, rodar
# o script "a seco" (como manda o DEPLOY.md) falhava com "DATABASE_URL não
# configurada", e a alternativa (source .env no shell) quebra no "&" dos
# parâmetros da URL do Neon. Descoberto na primeira carga real, 08/08/2026.
from dotenv import load_dotenv  # noqa: E402

load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"))

from core import bd, config, importacao, repositorio  # noqa: E402

# [ano, mês, cartão, saldo anterior, despesas, encargos, créditos, pagamentos, total]
# Confere exato: saldo + despesas + encargos − créditos − pagamentos = total.
RESUMOS_2026 = [
    (2026, 1, "Cartão Santander", 12082.51, 6621.11, 470.86, 0.00, 12553.37, 6621.11),
    (2026, 2, "Cartão Santander", 6621.11, 6114.64, 431.88, 0.00, 7133.05, 6034.58),
    (2026, 4, "Cartão Santander", 5295.05, 3730.16, 639.49, 0.00, 5296.00, 4368.70),
    (2026, 5, "Cartão Santander", 4368.70, 8334.77, 275.43, 11.75, 4369.00, 8598.15),
    (2026, 6, "Cartão Santander", 8598.15, 7731.91, 1185.41, 0.04, 4500.00, 13015.43),
    (2026, 7, "Cartão Santander", 13015.43, 8477.08, 739.20, 0.02, 8500.00, 13731.69),
    (2026, 8, "Cartão Santander", 13731.69, 6580.46, 494.79, 32.60, 14193.96, 6580.38),
]

COLUNAS = {"banco": 1, "fonte": 2, "data": 3, "descricao": 4, "categoria": 5,
           "subcategoria": 6, "item_fixo": 7, "conta": 8, "tipo": 9, "valor": 10,
           "competencia": 11, "status": 12, "arquivo": 13}


def ler_planilha(caminho: str):
    import openpyxl

    wb = openpyxl.load_workbook(caminho, data_only=True)
    if "Lançamentos" not in wb.sheetnames:
        sys.exit("A planilha não tem a aba 'Lançamentos'.")
    ws = wb["Lançamentos"]
    linhas = []
    for r in range(4, ws.max_row + 1):
        def celula(nome):
            return ws.cell(r, COLUNAS[nome]).value

        data, comp = celula("data"), celula("competencia")
        # Depois da última linha real vêm linhas de fórmula: competência vira
        # string em vez de data. É o corte natural.
        if not isinstance(data, dt.datetime) or not isinstance(comp, dt.datetime):
            continue
        valor = celula("valor")
        if valor is None:
            continue
        linhas.append({
            "data": data.date(), "competencia": dt.date(comp.year, comp.month, 1),
            "descricao": str(celula("descricao") or "")[:importacao.LIMITE_DESCRICAO],
            "categoria": celula("categoria") or "", "subcategoria": celula("subcategoria") or "",
            "item_fixo": celula("item_fixo") or "", "tipo": celula("tipo") or "",
            "conta": celula("conta") or "", "fonte": celula("fonte") or "",
            "valor": round(float(valor), 2), "status": celula("status") or config.STATUS_PADRAO,
            "arquivo": celula("arquivo") or "carga inicial",
        })
    wb.close()
    return linhas


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit("Uso: python3 scripts/migrar_da_planilha.py <caminho do .xlsx>")
    caminho = sys.argv[1]
    if not os.path.isfile(caminho):
        sys.exit("Arquivo não encontrado: %s" % caminho)
    if not bd.banco_configurado():
        sys.exit("DATABASE_URL não configurada.")

    contas = repositorio.mapa_contas()
    if not contas:
        sys.exit("Nenhuma conta cadastrada. Rode antes: python3 scripts/criar_schema.py")

    linhas = ler_planilha(caminho)
    print("Lidos %d lançamentos da planilha." % len(linhas))

    import collections
    vistas: collections.Counter = collections.Counter()
    preparadas, sem_conta = [], collections.Counter()
    for linha in linhas:
        conta = contas.get(linha["conta"])
        if not conta:
            sem_conta[linha["conta"]] += 1
            continue
        k = importacao.chave(conta["id"], linha["competencia"], linha["data"],
                             linha["descricao"], linha["valor"])
        vistas[k] += 1
        preparadas.append({
            "conta_id": conta["id"], "fonte": linha["fonte"], "data": linha["data"],
            "descricao": linha["descricao"], "categoria": linha["categoria"],
            "subcategoria": linha["subcategoria"], "item_fixo": linha["item_fixo"],
            "tipo": linha["tipo"], "valor": linha["valor"],
            "competencia": linha["competencia"], "status": linha["status"],
            "arquivo": str(linha["arquivo"]), "chave": k, "ocorrencia": vistas[k],
        })

    for nome, quantas in sem_conta.items():
        print("  ⚠ conta desconhecida '%s': %d linha(s) fora." % (nome, quantas))

    repetidas = sum(n - 1 for n in vistas.values() if n > 1)
    if repetidas:
        print("  %d linha(s) são repetição legítima da mesma chave "
              "(cobranças idênticas no mesmo dia) — todas preservadas." % repetidas)

    inseridos = repositorio.inserir_lancamentos(preparadas)
    print("Inseridos: %d · já existiam: %d" % (inseridos, len(preparadas) - inseridos))

    for ano, mes, conta_nome, saldo, desp, enc, cred, pag, total in RESUMOS_2026:
        conta = contas.get(conta_nome)
        if not conta:
            continue
        repositorio.upsert_resumo({
            "conta_id": conta["id"], "competencia": dt.date(ano, mes, 1),
            "saldo_anterior": saldo, "despesas": desp, "encargos": enc,
            "creditos": cred, "pagamentos": pag, "total_informado": total,
            "arquivo": "Fatura_%02d%04d (PDF)" % (mes, ano)})
    print("Resumos de fatura gravados: %d" % len(RESUMOS_2026))

    print("\nCompetências no banco: %s" % ", ".join(
        c.strftime("%Y-%m") for c in repositorio.competencias()))


if __name__ == "__main__":
    main()
